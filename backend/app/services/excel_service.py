"""Excel import/export service utilities."""

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook


def create_excel_file(headers: list[str], data: list[dict[str, Any]]) -> BytesIO:
    """Create an Excel file from headers and data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Write headers
    ws.append(headers)
    
    # Write data rows
    for row in data:
        ws.append([row.get(header, "") for header in headers])
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def read_excel_file(file_content: bytes) -> list[dict[str, Any]]:
    """Read data from Excel file."""
    wb = load_workbook(filename=BytesIO(file_content))
    ws = wb.active
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    
    # Read data rows
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):  # Skip empty rows
            data.append(dict(zip(headers, row)))
    
    return data


